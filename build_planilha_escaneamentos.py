#!/usr/bin/env python3
"""Planilha de escaneamentos e avaliacoes da QT Pizza Bar.

Le cliques.json (lista de {id, garcom, momento, ua}, momento em America/Sao_Paulo,
formato YYYY-MM-DD HH:MM:SS[.ffffff]) e, se existir, avaliacoes.json e perfil.json,
e escreve QT_escaneamentos.xlsx.

A planilha e reconstruida por inteiro a cada execucao, nunca acrescentada.
Cliques do mesmo atendente no mesmo aparelho dentro de JANELA_S segundos contam
como um so em todas as abas, e ficam visiveis e marcados na aba Base.

Sem avaliacoes.json a rodada continua valendo: sai o relatorio de escaneamentos
como sempre saiu, e o JSON devolve regra_do_zero = "sem_avaliacoes".

PRECISAO: o momento precisa vir com microssegundos. Truncado em segundo, o corte
de repeticao acusa 19 descontados onde ha 16, porque intervalos reais de 10,4s
aparecem como 10s. O SELECT que alimenta cliques.json usa 'YYYY-MM-DD HH24:MI:SS.US'.
"""
import json, sys, os, datetime as dt, zoneinfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

JANELA_S = 10           # repeticao: mesmo atendente, mesmo aparelho, ate 10s
JANELA_MIN = 10         # atribuicao: avaliacao olha ate 10 min para tras
ABERTURA_H = 18         # scan antes disso e teste da equipe, nao cliente
POR_AVALIACAO = 2.0     # R$ por avaliacao publicada
BONUS_LOTE = 50.0       # R$ por lote fechado
TAMANHO_LOTE = 40

SAIDA = sys.argv[1] if len(sys.argv) > 1 else "QT_escaneamentos.xlsx"
ESTADO = sys.argv[2] if len(sys.argv) > 2 else "avaliacoes_estado.json"

PRETO, CINZA, BRANCO, FONTE = "1A1E1E", "A0A5A5", "EFECEC", "Helvetica"
F_TIT = Font(name=FONTE, size=16, bold=True, color=PRETO)
F_SUB = Font(name=FONTE, size=10, color="6E7373")
F_HDR = Font(name=FONTE, size=10, bold=True, color=BRANCO)
F_TXT = Font(name=FONTE, size=11, color=PRETO)
F_TOT = Font(name=FONTE, size=11, bold=True, color=PRETO)
F_NOTA = Font(name=FONTE, size=9, color="6E7373")
FILL_HDR = PatternFill("solid", fgColor=PRETO)
FILL_BG = PatternFill("solid", fgColor=BRANCO)
FILL_TOT = PatternFill("solid", fgColor="DCD9D9")
FILL_FLAG = PatternFill("solid", fgColor="D8D4D4")
FILL_EDIT = PatternFill("solid", fgColor="FFF3C4")
BORDA = Border(bottom=Side(style="thin", color=CINZA))
DIAS_PT = ["seg", "ter", "qua", "qui", "sex", "sab", "dom"]
DIAS_LONGO = ["segunda", "terca", "quarta", "quinta", "sexta", "sabado", "domingo"]

agora = dt.datetime.now(zoneinfo.ZoneInfo("America/Sao_Paulo"))
GERADO_EM = agora.strftime("%d/%m/%Y %H:%M") + " (America/Sao_Paulo)"


def aparelho(ua):
    if "iPhone" in ua: d = "iPhone"
    elif "iPad" in ua: d = "iPad"
    elif "Android" in ua: d = "Android"
    elif "Macintosh" in ua: d = "Mac"
    elif "Windows" in ua: d = "Windows"
    else: d = "Outro"
    if "SamsungBrowser" in ua: n = "Samsung Internet"
    elif "CriOS" in ua: n = "Chrome"
    elif "Chrome" in ua: n = "Chrome"
    elif "Safari" in ua: n = "Safari"
    else: n = "Outro"
    v = ""
    for tag in ("Version/", "CriOS/", "SamsungBrowser/", "Chrome/"):
        if tag in ua:
            v = ua.split(tag, 1)[1].split(" ")[0].split(".")[0]
            break
    return f"{d} / {n} {v}".strip()


def instante(txt):
    """Aceita com e sem microssegundos. Sem eles a dedup fica imprecisa, mas roda."""
    txt = txt.strip().replace("T", " ")
    return dt.datetime.strptime(txt, "%Y-%m-%d %H:%M:%S.%f" if "." in txt else "%Y-%m-%d %H:%M:%S")


def le_json(caminho, padrao):
    if not os.path.exists(caminho):
        return padrao
    try:
        with open(caminho, encoding="utf-8") as fh:
            conteudo = json.load(fh)
    except (json.JSONDecodeError, OSError):
        return padrao
    return conteudo if conteudo else padrao


# ---------------------------------------------------------------- escaneamentos
with open("cliques.json", encoding="utf-8") as fh:
    bruto = json.load(fh)

reg = []
for r in bruto:
    ts = instante(r["momento"])
    reg.append({"id": r["id"], "garcom": r["garcom"], "ts": ts, "dia": ts.date(),
                "ua": r.get("ua", ""), "disp": aparelho(r.get("ua", ""))})
reg.sort(key=lambda r: (r["ts"], r["id"]))

# Compara sempre com a linha anterior BRUTA do mesmo par (atendente, aparelho).
# Comparar com o ultimo mantido preservaria o terceiro toque de um toque triplo.
ultimo = {}
for r in reg:
    chave = (r["garcom"], r["ua"])
    ant = ultimo.get(chave)
    gap = (r["ts"] - ant).total_seconds() if ant else None
    r["gap"] = gap
    r["descontado"] = bool(ant and gap is not None and gap <= JANELA_S)
    r["teste"] = r["ts"].hour < ABERTURA_H
    ultimo[chave] = r["ts"]

validos = [r for r in reg if not r["descontado"]]
n_desc = len(reg) - len(validos)
n_teste = sum(1 for r in validos if r["teste"])

dias = sorted({r["dia"] for r in validos}) or sorted({r["dia"] for r in reg})
atendentes = sorted({r["garcom"] for r in reg})

# ------------------------------------------------------------------ avaliacoes
cru = le_json("avaliacoes.json", [])
perfil = le_json("perfil.json", {})

avaliacoes = []
for a in cru:
    ts = instante(a["criado_em"])
    avaliacoes.append({
        "review_id": a.get("review_id") or a.get("reviewId") or "",
        "nota": int(a["nota"]),
        "ts": ts, "dia": ts.date(),
        "atualizado_em": a.get("atualizado_em") or a.get("criado_em"),
        "cliente": a.get("cliente") or "",
        "tem_texto": bool(a.get("tem_texto")),
        "nota_anterior": a.get("nota_anterior"),
    })
avaliacoes.sort(key=lambda a: (a["ts"], a["review_id"]))
tem_avaliacoes = bool(avaliacoes)

# Pareamento 1 para 1 sobre a lista ja liquida de repeticoes.
# Cada avaliacao leva o escaneamento livre mais proximo dentro de [t-10min, t].
# Sem candidato a avaliacao e orfa; escaneamento nao consumido e perdido.
livres = sorted(validos, key=lambda r: r["ts"])
consumidos = set()
pares, orfas = [], []
for a in avaliacoes:
    inicio = a["ts"] - dt.timedelta(minutes=JANELA_MIN)
    melhor = None
    for r in livres:
        if r["ts"] > a["ts"]:
            break
        if r["id"] in consumidos or r["ts"] < inicio:
            continue
        if melhor is None or r["ts"] > melhor["ts"]:
            melhor = r
    if melhor is None:
        orfas.append(a)
        continue
    consumidos.add(melhor["id"])
    melhor["par"] = a
    a["par"] = melhor
    pares.append((melhor, a))

perdidos = [r for r in validos if r["id"] not in consumidos]

# Nota editada: o cliente muda a nota depois que a avaliacao ja foi paga.
# Duas fontes, porque a tarefa diaria roda em container efemero: a coluna
# nota_anterior que a sincronizacao grava no banco, e um estado local, util
# quando o script roda em host que persiste arquivo.
anterior = {}
if os.path.exists(ESTADO):
    for k, v in (le_json(ESTADO, {}) or {}).items():
        anterior[k] = v
notas_alteradas = []
for a in avaliacoes:
    de = a["nota_anterior"]
    if de is None:
        guardado = anterior.get(a["review_id"])
        if guardado and guardado.get("nota") is not None and int(guardado["nota"]) != a["nota"]:
            de = int(guardado["nota"])
    if de is not None and int(de) != a["nota"]:
        notas_alteradas.append({"review_id": a["review_id"], "cliente": a["cliente"],
                                "de": int(de), "para": a["nota"],
                                "quando": a["atualizado_em"]})
if tem_avaliacoes:
    try:
        with open(ESTADO, "w", encoding="utf-8") as fh:
            json.dump({a["review_id"]: {"nota": a["nota"], "atualizado_em": a["atualizado_em"]}
                       for a in avaliacoes}, fh, ensure_ascii=False)
    except OSError:
        pass

# ---------------------------------------------------------------- agregacoes
agg = {}
for r in validos:
    a = agg.setdefault((r["dia"], r["garcom"]),
                       {"n": 0, "notas": 0, "min": r["ts"], "max": r["ts"]})
    a["n"] += 1
    if r.get("par"): a["notas"] += 1
    a["min"] = min(a["min"], r["ts"])
    a["max"] = max(a["max"], r["ts"])

wb = Workbook()


def cabecalho(ws, titulo, sub, largura, linhas=0):
    ws.sheet_view.showGridLines = False
    ws["A1"] = "QT PIZZA BAR"
    ws["A1"].font = Font(name=FONTE, size=9, bold=True, color=CINZA)
    ws["A2"] = titulo; ws["A2"].font = F_TIT
    ws["A3"] = sub; ws["A3"].font = F_SUB
    for row in ws.iter_rows(min_row=1, max_row=max(60, linhas or len(reg) + 20),
                            min_col=1, max_col=largura):
        for c in row:
            c.fill = FILL_BG
    ws.row_dimensions[2].height = 24


def faixa(ws, linha, valores, larguras):
    for i, (v, w) in enumerate(zip(valores, larguras), start=1):
        c = ws.cell(row=linha, column=i, value=v)
        c.font = F_HDR; c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center" if i > 1 else "left",
                                vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[linha].height = 26


# ---- Diario
ws = wb.active; ws.title = "Diario"
cabecalho(ws, "Escaneamentos e avaliacoes por dia e atendente",
          "Formato longo, ja liquido de repeticoes. Base para tabela dinamica.", 10)
faixa(ws, 5, ["Data", "Dia", "Atendente", "Escaneamentos", "Viraram nota", "Perdidos",
              "Conversao", "Media das notas", "Primeiro", "Ultimo"],
      [13, 7, 18, 15, 14, 11, 12, 15, 11, 11])
linha = 6
PRIM_DIARIO = 6
for d in dias:
    for g in atendentes:
        a = agg.get((d, g))
        ws.cell(row=linha, column=1, value=d).number_format = "DD/MM/YYYY"
        ws.cell(row=linha, column=2, value=DIAS_PT[d.weekday()])
        ws.cell(row=linha, column=3, value=g)
        ws.cell(row=linha, column=4, value=a["n"] if a else 0)
        ws.cell(row=linha, column=5, value=a["notas"] if a else 0)
        ws.cell(row=linha, column=6, value=f"=D{linha}-E{linha}")
        ws.cell(row=linha, column=7, value=f"=IFERROR(E{linha}/D{linha},0)").number_format = "0.0%"
        ws.cell(row=linha, column=8, value="")   # preenchida adiante, depende do Pareamento
        ws.cell(row=linha, column=9, value=a["min"].strftime("%H:%M:%S") if a else "")
        ws.cell(row=linha, column=10, value=a["max"].strftime("%H:%M:%S") if a else "")
        for col in range(1, 11):
            c = ws.cell(row=linha, column=col); c.font = F_TXT; c.border = BORDA
            if col >= 4: c.alignment = Alignment(horizontal="center")
        linha += 1
FIM = linha - 1
ws.freeze_panes = "A6"
ws.cell(row=linha + 1, column=1,
        value=f"Gerado em {GERADO_EM}. Zero significa nenhum QR lido naquele dia, "
              "nao ausencia do turno.").font = F_NOTA
ws.cell(row=linha + 2, column=1,
        value=f"Viraram nota: escaneamentos pareados 1 para 1 com avaliacao publicada em ate "
              f"{JANELA_MIN} minutos. Perdidos: o resto.").font = F_NOTA

RD, RA, RV = f"Diario!$A${PRIM_DIARIO}:$A${FIM}", f"Diario!$C${PRIM_DIARIO}:$C${FIM}", f"Diario!$D${PRIM_DIARIO}:$D${FIM}"
RN = f"Diario!$E${PRIM_DIARIO}:$E${FIM}"

# ---- Matriz
ws = wb.create_sheet("Matriz")
cabecalho(ws, "Matriz dia x atendente",
          "Leitura rapida da semana. Todas as celulas sao formulas sobre a aba Diario.",
          len(atendentes) + 2)
faixa(ws, 5, ["Data"] + atendentes + ["Total da casa"],
      [13] + [16] * len(atendentes) + [16])
linha = 6
for d in dias:
    ws.cell(row=linha, column=1, value=d).number_format = "DD/MM/YYYY"
    for j, g in enumerate(atendentes, start=2):
        L = get_column_letter(j)
        ws.cell(row=linha, column=j, value=f"=SUMIFS({RV},{RD},$A{linha},{RA},{L}$5)")
    ult = get_column_letter(len(atendentes) + 1)
    ws.cell(row=linha, column=len(atendentes) + 2, value=f"=SUM(B{linha}:{ult}{linha})")
    for col in range(1, len(atendentes) + 3):
        c = ws.cell(row=linha, column=col)
        c.font = F_TOT if col == len(atendentes) + 2 else F_TXT
        c.border = BORDA
        if col > 1: c.alignment = Alignment(horizontal="center"); c.number_format = "0"
    linha += 1
for col in range(1, len(atendentes) + 3):
    L = get_column_letter(col)
    c = ws.cell(row=linha, column=col,
                value="Total" if col == 1 else f"=SUM({L}6:{L}{linha-1})")
    c.font = F_TOT; c.fill = FILL_TOT
    if col > 1: c.alignment = Alignment(horizontal="center"); c.number_format = "0"
TOT_LINHA = linha
TOT_COL = get_column_letter(len(atendentes) + 2)
ws.freeze_panes = "B6"

# ---- Pareamento  (montada antes de Atendentes: as medias sao formulas sobre ela)
ws_par = wb.create_sheet("Pareamento")
PAR_PRIM = 6
PAR_ULT = max(PAR_PRIM, PAR_PRIM + len(pares) - 1)
RP_AT = f"Pareamento!$B${PAR_PRIM}:$B${PAR_ULT}"
RP_DT = f"Pareamento!$C${PAR_PRIM}:$C${PAR_ULT}"
RP_NT = f"Pareamento!$G${PAR_PRIM}:$G${PAR_ULT}"

# ---- Atendentes
ws = wb.create_sheet("Atendentes")
cabecalho(ws, "Consolidado por atendente",
          "Acumulado do periodo. Participacao e sobre o total da casa.", 10)
faixa(ws, 5, ["Atendente", "Escaneamentos", "Viraram nota", "Perdidos", "Conversao",
              "Media das notas", "Dias com leitura", "Media por dia com leitura",
              "Participacao", "Ultimo escaneamento"],
      [18, 15, 14, 11, 12, 15, 15, 20, 14, 20])
ordem = sorted(atendentes, key=lambda g: -sum(a["n"] for (d, gg), a in agg.items() if gg == g))
linha = 6
for g in ordem:
    ults = [a["max"] for (d, gg), a in agg.items() if gg == g]
    ws.cell(row=linha, column=1, value=g)
    ws.cell(row=linha, column=2, value=f"=SUMIF({RA},$A{linha},{RV})")
    ws.cell(row=linha, column=3, value=f"=SUMIF({RA},$A{linha},{RN})")
    ws.cell(row=linha, column=4, value=f"=B{linha}-C{linha}")
    ws.cell(row=linha, column=5, value=f"=IFERROR(C{linha}/B{linha},0)")
    ws.cell(row=linha, column=6, value=f"=IFERROR(AVERAGEIF({RP_AT},$A{linha},{RP_NT}),0)")
    ws.cell(row=linha, column=7, value=f'=COUNTIFS({RA},$A{linha},{RV},">0")')
    ws.cell(row=linha, column=8, value=f"=IFERROR(B{linha}/G{linha},0)")
    ws.cell(row=linha, column=9, value=f"=IFERROR(B{linha}/Matriz!${TOT_COL}${TOT_LINHA},0)")
    ws.cell(row=linha, column=10,
            value=max(ults).strftime("%d/%m %H:%M") if ults else "sem registro")
    for col in range(1, 11):
        c = ws.cell(row=linha, column=col); c.font = F_TXT; c.border = BORDA
        if col > 1: c.alignment = Alignment(horizontal="center")
    for col, fmt in ((2, "0"), (3, "0"), (4, "0"), (5, "0.0%"), (6, "0.00"),
                     (7, "0"), (8, "0.0"), (9, "0.0%")):
        ws.cell(row=linha, column=col).number_format = fmt
    linha += 1
for col, val in ((1, "Casa"), (2, f"=SUM(B6:B{linha-1})"), (3, f"=SUM(C6:C{linha-1})"),
                 (4, f"=SUM(D6:D{linha-1})"), (5, f"=IFERROR(C{linha}/B{linha},0)"),
                 (6, f"=IFERROR(AVERAGE({RP_NT}),0)"),
                 (9, f"=IFERROR(SUM(I6:I{linha-1}),0)")):
    ws.cell(row=linha, column=col, value=val)
for col in range(1, 11):
    c = ws.cell(row=linha, column=col); c.font = F_TOT; c.fill = FILL_TOT
    if col > 1: c.alignment = Alignment(horizontal="center")
for col, fmt in ((2, "0"), (3, "0"), (4, "0"), (5, "0.0%"), (6, "0.00"), (9, "0.0%")):
    ws.cell(row=linha, column=col).number_format = fmt
ws.cell(row=linha + 2, column=1,
        value=f"Numeros liquidos. {len(reg)} cliques na base, {n_desc} descontados como repeticao, "
              f"{len(validos)} contados.").font = F_NOTA
ws.cell(row=linha + 3, column=1,
        value=("Media das notas so considera avaliacao pareada. Avaliacao orfa nao entra em "
               "nenhum atendente.")).font = F_NOTA

# ---- Pareamento (conteudo)
ws = ws_par
cabecalho(ws, "Pareamento escaneamento x avaliacao",
          f"Uma linha por par. Janela de {JANELA_MIN} minutos, cada escaneamento consumido "
          "uma unica vez. Aba de auditoria do pagamento.", 9, linhas=len(pares) + 30)
faixa(ws, 5, ["#", "Atendente", "Data", "Hora do scan", "Hora da avaliacao",
              "Defasagem (min)", "Nota", "Cliente", "Tem texto"],
      [6, 16, 13, 14, 16, 15, 8, 26, 11])
linha = PAR_PRIM
for i, (esc, av) in enumerate(pares, start=1):
    dif = (av["ts"] - esc["ts"]).total_seconds() / 60
    ws.cell(row=linha, column=1, value=i)
    ws.cell(row=linha, column=2, value=esc["garcom"])
    ws.cell(row=linha, column=3, value=esc["dia"]).number_format = "DD/MM/YYYY"
    ws.cell(row=linha, column=4, value=esc["ts"].strftime("%H:%M:%S"))
    ws.cell(row=linha, column=5, value=av["ts"].strftime("%H:%M:%S"))
    ws.cell(row=linha, column=6, value=round(dif, 2)).number_format = "0.00"
    ws.cell(row=linha, column=7, value=av["nota"])
    ws.cell(row=linha, column=8, value=av["cliente"])
    ws.cell(row=linha, column=9, value="sim" if av["tem_texto"] else "nao")
    for col in range(1, 10):
        c = ws.cell(row=linha, column=col); c.font = F_TXT; c.border = BORDA
        if col != 8: c.alignment = Alignment(horizontal="center")
    linha += 1
if not pares:                      # mantem o intervalo valido na rodada degradada
    for col in range(1, 10):
        ws.cell(row=linha, column=col, value="").border = BORDA
    linha += 1
LIN_TOT_PAR = linha
ws.cell(row=linha, column=1, value="TOTAL").font = F_TOT
ws.cell(row=linha, column=2, value=f"=COUNTA(B{PAR_PRIM}:B{PAR_ULT})")
ws.cell(row=linha, column=6, value=f"=IFERROR(AVERAGE(F{PAR_PRIM}:F{PAR_ULT}),0)").number_format = "0.00"
ws.cell(row=linha, column=7, value=f"=IFERROR(AVERAGE(G{PAR_PRIM}:G{PAR_ULT}),0)").number_format = "0.00"
for col in range(1, 10):
    c = ws.cell(row=linha, column=col); c.font = F_TOT; c.fill = FILL_TOT
    if col != 8: c.alignment = Alignment(horizontal="center")
ws.freeze_panes = f"A{PAR_PRIM}"

linha += 2
faixa(ws, linha, ["#", "Data", "Hora", "Nota", "Cliente", "Tem texto", "", "", ""],
      [6, 13, 14, 8, 26, 11, 2, 2, 2])
ws.cell(row=linha - 1, column=1,
        value="Avaliacoes sem escaneamento correspondente (orfas). Nao sao creditadas a "
              "ninguem e nao entram no pagamento.").font = F_NOTA
linha += 1
for i, av in enumerate(orfas, start=1):
    ws.cell(row=linha, column=1, value=i)
    ws.cell(row=linha, column=2, value=av["dia"]).number_format = "DD/MM/YYYY"
    ws.cell(row=linha, column=3, value=av["ts"].strftime("%H:%M:%S"))
    ws.cell(row=linha, column=4, value=av["nota"])
    ws.cell(row=linha, column=5, value=av["cliente"])
    ws.cell(row=linha, column=6, value="sim" if av["tem_texto"] else "nao")
    for col in range(1, 7):
        c = ws.cell(row=linha, column=col); c.font = F_TXT; c.border = BORDA
        if col != 5: c.alignment = Alignment(horizontal="center")
    linha += 1
ws.cell(row=linha + 1, column=1,
        value=f"{len(pares)} pares, {len(orfas)} orfas, {len(perdidos)} escaneamentos perdidos. "
              "Um escaneamento so pode ser consumido uma vez.").font = F_NOTA

# ---- Pagamento
ws = wb.create_sheet("Pagamento")
cabecalho(ws, "Pagamento por atendente",
          "Parametros nas celulas amarelas. Todo o calculo e formula, nenhum valor fixo.",
          14, linhas=40)
faixa(ws, 5, ["Atendente", "Escaneamentos", "Viraram nota", "Perderam", "Conversao",
              "Media das notas", "Variavel", "Bonus", "Total a pagar",
              "Faltam p/ proximo lote"],
      [18, 15, 14, 11, 12, 15, 13, 11, 14, 20])
ws.column_dimensions["K"].width = 3
ws.column_dimensions["L"].width = 3
ws.column_dimensions["M"].width = 24
ws.column_dimensions["N"].width = 10
par = [("Parametros", None), ("Valor por avaliacao", POR_AVALIACAO),
       ("Bonus por lote", BONUS_LOTE), ("Avaliacoes por lote", TAMANHO_LOTE)]
for i, (rot, val) in enumerate(par):
    c = ws.cell(row=5 + i, column=13, value=rot)
    c.font = F_HDR if val is None else F_TXT
    if val is None: c.fill = FILL_HDR
    if val is not None:
        v = ws.cell(row=5 + i, column=14, value=val)
        v.font = F_TOT; v.fill = FILL_EDIT; v.alignment = Alignment(horizontal="center")
        v.number_format = "0.00" if i < 3 else "0"
PG_PRIM = 6
linha = PG_PRIM
for g in ordem:
    ws.cell(row=linha, column=1, value=g)
    ws.cell(row=linha, column=2, value=f"=SUMIF({RA},$A{linha},{RV})")
    ws.cell(row=linha, column=3, value=f"=SUMIF({RA},$A{linha},{RN})")
    ws.cell(row=linha, column=4, value=f"=B{linha}-C{linha}")
    ws.cell(row=linha, column=5, value=f"=IFERROR(C{linha}/B{linha},0)")
    ws.cell(row=linha, column=6, value=f"=IFERROR(AVERAGEIF({RP_AT},$A{linha},{RP_NT}),0)")
    ws.cell(row=linha, column=7, value=f"=C{linha}*$N$6")
    ws.cell(row=linha, column=8, value=f"=IFERROR(INT(C{linha}/$N$8)*$N$7,0)")
    ws.cell(row=linha, column=9, value=f"=G{linha}+H{linha}")
    ws.cell(row=linha, column=10, value=f"=IFERROR($N$8-MOD(C{linha},$N$8),0)")
    for col in range(1, 11):
        c = ws.cell(row=linha, column=col); c.font = F_TXT; c.border = BORDA
        if col > 1: c.alignment = Alignment(horizontal="center")
    for col, fmt in ((2, "0"), (3, "0"), (4, "0"), (5, "0.0%"), (6, "0.00"),
                     (7, '"R$" #,##0.00'), (8, '"R$" #,##0.00'),
                     (9, '"R$" #,##0.00'), (10, "0")):
        ws.cell(row=linha, column=col).number_format = fmt
    linha += 1
PG_ULT = linha - 1
ws.cell(row=linha, column=1, value="TOTAL")
for col in (2, 3, 4, 7, 8, 9):
    L = get_column_letter(col)
    ws.cell(row=linha, column=col, value=f"=SUM({L}{PG_PRIM}:{L}{PG_ULT})")
ws.cell(row=linha, column=5, value=f"=IFERROR(C{linha}/B{linha},0)")
ws.cell(row=linha, column=6, value=f"=IFERROR(AVERAGE({RP_NT}),0)")
for col in range(1, 11):
    c = ws.cell(row=linha, column=col); c.font = F_TOT; c.fill = FILL_TOT
    if col > 1: c.alignment = Alignment(horizontal="center")
for col, fmt in ((2, "0"), (3, "0"), (4, "0"), (5, "0.0%"), (6, "0.00"),
                 (7, '"R$" #,##0.00'), (8, '"R$" #,##0.00'), (9, '"R$" #,##0.00')):
    ws.cell(row=linha, column=col).number_format = fmt
LIN_TOT_PG = linha
ws.cell(row=linha + 2, column=1, value="Custo por avaliacao publicada")
ws.cell(row=linha + 2, column=1).font = F_TOT
c = ws.cell(row=linha + 2, column=3,
            value=f"=IFERROR(I{LIN_TOT_PG}/C{LIN_TOT_PG},0)")
c.font = F_TOT; c.number_format = '"R$" #,##0.00'; c.alignment = Alignment(horizontal="center")
for i, txt in enumerate([
    "So entra no calculo avaliacao pareada 1 para 1. Orfa nao e creditada a ninguem.",
    f"Lote incompleto nao gera bonus, nao acumula entre atendentes nem entre meses.",
    "Mudou o valor por avaliacao ou o tamanho do lote, altere so as celulas amarelas em M/N.",
]):
    ws.cell(row=linha + 4 + i, column=1, value=txt).font = F_NOTA

# ---- media das notas no Diario (depende das faixas do Pareamento)
wsd = wb["Diario"]
for lin in range(PRIM_DIARIO, FIM + 1):
    wsd.cell(row=lin, column=8,
             value=f"=IFERROR(AVERAGEIFS({RP_NT},{RP_AT},$C{lin},{RP_DT},$A{lin}),0)")
    wsd.cell(row=lin, column=8).number_format = "0.00"
    wsd.cell(row=lin, column=8).alignment = Alignment(horizontal="center")
    wsd.cell(row=lin, column=8).font = F_TXT
    wsd.cell(row=lin, column=8).border = BORDA

# ---- Base
ws = wb.create_sheet("Base")
cabecalho(ws, "Base de cliques",
          "Todos os registros, inclusive os descontados. Use para auditar um numero estranho.", 7)
faixa(ws, 5, ["ID", "Data e hora", "Atendente", "Aparelho e navegador",
              "Intervalo do anterior (s)", "Descontado", "Situacao"],
      [7, 22, 16, 30, 22, 13, 14])
linha = 6
for r in reg:
    if r["descontado"]: sit = "repeticao"
    elif r.get("par"): sit = "virou nota"
    elif r["teste"]: sit = "teste"
    else: sit = "perdido"
    ws.cell(row=linha, column=1, value=r["id"])
    ws.cell(row=linha, column=2, value=r["ts"]).number_format = "DD/MM/YYYY HH:MM:SS"
    ws.cell(row=linha, column=3, value=r["garcom"])
    ws.cell(row=linha, column=4, value=r["disp"])
    c = ws.cell(row=linha, column=5, value=round(r["gap"], 3) if r["gap"] is not None else "")
    c.number_format = "0.000"
    ws.cell(row=linha, column=6, value="sim" if r["descontado"] else "")
    ws.cell(row=linha, column=7, value=sit)
    for col in range(1, 8):
        c = ws.cell(row=linha, column=col)
        c.font = F_TOT if r["descontado"] else F_TXT
        c.border = BORDA
        if col in (1, 5, 6, 7): c.alignment = Alignment(horizontal="center")
        if r["descontado"] or r["teste"]: c.fill = FILL_FLAG
    linha += 1
ws.freeze_panes = "A6"
ws.cell(row=linha + 1, column=1,
        value=f"Intervalo compara com o clique anterior do MESMO atendente no MESMO aparelho. "
              f"Descontado quando esse intervalo e de ate {JANELA_S} segundos.").font = F_NOTA
ws.cell(row=linha + 2, column=1,
        value=f"Situacao teste: escaneamento antes das {ABERTURA_H}h, fora do horario de servico. "
              f"{n_teste} no periodo.").font = F_NOTA

# ---- Leitura
ws = wb.create_sheet("Leitura")
cabecalho(ws, "Como ler esta planilha", "Definicoes, premissas e limites do dado.", 2, linhas=40)
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 96
itens = [
    ("Escaneamento", "Um QR code lido, nao uma avaliacao publicada no Google. Sao numeros "
     "diferentes: o cliente pode ler o codigo e nao escrever nada. Nunca troque os dois termos "
     "ao apresentar para a equipe."),
    ("Avaliacao publicada", "Avaliacao que apareceu no perfil do Google. E o que gera pagamento, "
     "e so quando pareada com um escaneamento."),
    ("Nota da casa", "Media exibida no Google sobre toda a base historica, centenas de avaliacoes "
     "antigas. Move devagar e nao e resultado do dia."),
    ("Fonte", "Tabela public.cliques_avaliacao do projeto Supabase helinoirdizwrluydkzp, e "
     "public.avaliacoes para as notas. Todos os cortes de dia sao no fuso America/Sao_Paulo, "
     "entao um QR lido as 22h30 de sabado fica no sabado."),
    ("Como e atualizada", "Reconstruida por inteiro a cada geracao, nunca acrescentada linha a "
     "linha. Rodada perdida se resolve sozinha na proxima, e gerar duas vezes da o mesmo resultado."),
    ("Desconto de repeticao", f"Dois cliques do mesmo atendente no mesmo aparelho dentro de "
     f"{JANELA_S} segundos contam como um. E o padrao de toque duplo na tela, que registraria duas "
     "leituras para um cliente so. O aparelho entra na chave porque dois clientes na mesma mesa nao "
     "podem ser fundidos. Todas as abas mostram o numero liquido; a aba Base mostra tudo."),
    ("Pareamento", f"Cada avaliacao leva o escaneamento livre mais proximo dentro dos "
     f"{JANELA_MIN} minutos anteriores, e esse escaneamento nao pode ser usado de novo. A janela foi "
     "medida, nao arbitrada: mediana de 54 segundos entre o escaneamento e o envio, p95 de 6,4 "
     "minutos, e nenhuma ocorrencia entre 20 e 60 minutos."),
    ("Orfa e perdido", "Orfa e avaliacao sem escaneamento livre na janela, provavelmente organica. "
     "Perdido e escaneamento que nao virou avaliacao. Orfa nao e creditada a ninguem."),
    ("Teste", f"Escaneamento antes das {ABERTURA_H}h e teste da equipe, nao cliente. Fica marcado "
     "na aba Base para poder sair da conta quando se quiser a conversao so de cliente."),
    ("Nota que muda", "O cliente pode editar a avaliacao depois. A rodada compara a nota de cada "
     "avaliacao com a da rodada anterior e avisa quando muda, porque uma nota 5 que vira 2 depois "
     "do pagamento passaria despercebida."),
    ("Zero", "Zero significa nenhum QR lido naquele dia por aquela pessoa. Nao significa falta. "
     "Quem entrou na equipe depois aparece com zero nos dias anteriores."),
    ("Nao atribua nota baixa", "O pareamento serve para creditar, nao para culpar. Cliente "
     "insatisfeito costuma avaliar fora da janela, entao atribuir negativa a uma pessoa e "
     "estatisticamente fragil e inverte o incentivo nas mesas de risco."),
    ("Formulas", "Matriz, Atendentes e Pagamento sao formulas sobre Diario e Pareamento. Corrigiu "
     "no Diario, o resto acompanha. As abas Base e Pareamento sao registro e nao devem ser editadas."),
]
linha = 5
for t, txt in itens:
    a = ws.cell(row=linha, column=1, value=t)
    a.font = Font(name=FONTE, size=11, bold=True, color=PRETO)
    a.alignment = Alignment(vertical="top")
    b = ws.cell(row=linha, column=2, value=txt)
    b.font = Font(name=FONTE, size=10, color="3A3F3F")
    b.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[linha].height = 48
    linha += 1
ws.cell(row=linha + 1, column=1, value=f"Gerado em {GERADO_EM}").font = F_NOTA

wb.save(SAIDA)

# ---- resumo para o relatorio de texto, na MESMA regra de desconto da planilha
hoje = agora.date()
ontem = hoje - dt.timedelta(days=1)
inicio_mes = hoje.replace(day=1)


def conta(filtro, g=None):
    return sum(1 for r in validos if filtro(r["dia"]) and (g is None or r["garcom"] == g))


def conta_notas(filtro, g=None):
    return sum(1 for r in validos
               if r.get("par") and filtro(r["dia"]) and (g is None or r["garcom"] == g))


por_atendente = sorted(((g, conta(lambda d: d == ontem, g), conta_notas(lambda d: d == ontem, g))
                        for g in atendentes), key=lambda t: (-t[1], t[0]))
totais = {
    "hoje": conta(lambda d: d == hoje),
    "ontem": conta(lambda d: d == ontem),
    "ultimos_7d": conta(lambda d: d >= hoje - dt.timedelta(days=6)),
    "mes": conta(lambda d: d >= inicio_mes),
    "total": len(validos),
}
ult = max(validos, key=lambda r: r["ts"]) if validos else None

limpos_ontem = totais["ontem"]
notas_ontem = conta_notas(lambda d: d == ontem)
teste_ontem = sum(1 for r in validos if r["dia"] == ontem and r["teste"])
notas_do_dia = [a["nota"] for esc, a in pares if esc["dia"] == ontem]

if not reg:
    regra = "vazio"
elif totais["total"] == totais["hoje"]:
    regra = "sem_historico"
elif totais["ontem"] == 0 and ontem.weekday() == 0:
    regra = "segunda_fechada"
elif totais["ontem"] == 0:
    regra = "zero_operacao"
elif not tem_avaliacoes:
    regra = "sem_avaliacoes"
else:
    regra = "normal"


def taxa(n, d):
    return round(n / d, 4) if d else 0.0


print(json.dumps({
    "arquivo": SAIDA,
    "gerado_em": GERADO_EM,
    "cliques_brutos": len(reg), "descontados": n_desc, "contados": len(validos),
    "periodo": {
        "escaneamentos_limpos": len(validos),
        "scans_teste": n_teste,
        "avaliacoes": len(avaliacoes),
        "viraram_nota": len(pares),
        "orfas": len(orfas),
        "perdidos": len(perdidos),
        "conversao": taxa(len(pares), len(validos)),
        "conversao_sem_teste": taxa(len(pares), len(validos) - n_teste),
        "media_notas": round(sum(a["nota"] for _, a in pares) / len(pares), 2) if pares else None,
        "defasagem_media_min": round(
            sum((a["ts"] - e["ts"]).total_seconds() for e, a in pares) / len(pares) / 60, 2)
        if pares else None,
    },
    "relatorio": {
        "data_ontem": ontem.strftime("%d/%m"),
        "dia_semana_ontem": DIAS_LONGO[ontem.weekday()],
        "ontem_por_atendente": [{"garcom": g, "escaneamentos": n, "viraram_nota": v}
                                for g, n, v in por_atendente],
        "totais_da_casa": totais,
        "conversao_ontem": taxa(notas_ontem, limpos_ontem),
        "viraram_nota_ontem": notas_ontem,
        "scans_teste_ontem": teste_ontem,
        "media_notas_ontem": round(sum(notas_do_dia) / len(notas_do_dia), 2) if notas_do_dia else None,
        "orfas_ontem": sum(1 for a in orfas if a["dia"] == ontem),
        "notas_alteradas": notas_alteradas,
        "perfil_google": {
            "nota_exibida": perfil.get("nota_exibida"),
            "total_avaliacoes": perfil.get("total_avaliacoes"),
        },
        "ultimo_escaneamento": ({"garcom": ult["garcom"], "quando": ult["ts"].strftime("%d/%m %H:%M")}
                                if ult else None),
        "regra_do_zero": regra,
        "avaliacoes_ausentes": not tem_avaliacoes,
    },
}, ensure_ascii=False))
