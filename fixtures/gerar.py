#!/usr/bin/env python3
"""Monta cliques.json e avaliacoes.json de agosto/2026 para validar o build.

cliques.json  vem de cliques_raw.csv + ua_map.txt (extraidos de cliques_avaliacao).
avaliacoes.json vem das abas da QT-scans-x-avaliacoes-agosto.xlsx, que e o
resultado ja conferido do pareamento: 112 pareadas + 14 orfas = 126 avaliacoes.
"""
import csv, json, hashlib, datetime as dt, sys, pathlib

AQUI = pathlib.Path(__file__).resolve().parent
RAIZ = AQUI.parent
INI, FIM = dt.date(2026, 8, 5), dt.date(2026, 8, 24)

uas = {}
for linha in (AQUI / "ua_map.txt").read_text(encoding="utf-8").splitlines():
    if linha.strip():
        i, ua = linha.split("\t", 1)
        uas[i] = ua

# Os microssegundos vao separados porque a dedup depende deles: truncar em
# segundo inventa repeticao onde o intervalo real passa de JANELA_S.
frac = dict(p.split(":") for p in (AQUI / "microssegundos.txt").read_text().split())

cliques = []
with open(AQUI / "cliques_raw.csv", encoding="utf-8") as fh:
    for cid, garcom, momento, idx in csv.reader(fh):
        if INI <= dt.datetime.strptime(momento, "%Y-%m-%d %H:%M:%S").date() <= FIM:
            cliques.append({"id": int(cid), "garcom": garcom,
                            "momento": f"{momento}.{frac[cid]}", "ua": uas[idx]})

import openpyxl
wb = openpyxl.load_workbook(RAIZ / "QT-scans-x-avaliacoes-agosto.xlsx")


def texto(v):
    if isinstance(v, dt.datetime): return v
    return str(v).strip() if v is not None else ""


def carimbo(data, hora):
    if isinstance(data, dt.datetime): d = data.date()
    else: d = dt.datetime.strptime(str(data).strip(), "%d/%m/%Y").date()
    h = hora.time() if isinstance(hora, dt.datetime) else dt.datetime.strptime(str(hora).strip(), "%H:%M:%S").time()
    return dt.datetime.combine(d, h)


avaliacoes = []
ws = wb["Scans que viraram nota"]
for r in ws.iter_rows(min_row=2, max_row=113, values_only=True):
    if r[0] is None or not isinstance(r[0], int): continue
    avaliacoes.append((carimbo(r[2], r[4]), int(r[6]), texto(r[7]), texto(r[8]) == "sim"))
ws = wb["Avaliacoes sem scan"]
for r in ws.iter_rows(min_row=2, max_row=15, values_only=True):
    if r[0] is None or not isinstance(r[0], int): continue
    avaliacoes.append((carimbo(r[1], r[2]), int(r[3]), texto(r[4]), texto(r[5]) == "sim"))

avaliacoes.sort(key=lambda a: (a[0], a[2]))
saida = []
for ts, nota, cliente, tem_texto in avaliacoes:
    rid = "fixture_" + hashlib.sha1(f"{ts.isoformat()}|{cliente}".encode()).hexdigest()[:16]
    saida.append({"review_id": rid, "nota": nota,
                  "criado_em": ts.strftime("%Y-%m-%d %H:%M:%S"),
                  "atualizado_em": ts.strftime("%Y-%m-%d %H:%M:%S"),
                  "cliente": cliente, "tem_texto": tem_texto, "respondida": False})

destino = pathlib.Path(sys.argv[1]) if len(sys.argv) > 1 else RAIZ
(destino / "cliques.json").write_text(json.dumps(cliques, ensure_ascii=False, indent=1), encoding="utf-8")
(destino / "avaliacoes.json").write_text(json.dumps(saida, ensure_ascii=False, indent=1), encoding="utf-8")
print(f"cliques.json: {len(cliques)} cliques brutos ({INI} a {FIM})")
print(f"avaliacoes.json: {len(saida)} avaliacoes")
