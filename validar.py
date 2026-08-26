#!/usr/bin/env python3
"""Confere o build contra os numeros medidos de agosto de 2026 (CLAUDE.md secao 8).

Roda o build sobre os fixtures e compara o JSON de saida e as celulas ja
recalculadas da planilha. Se algum numero divergir, sai com codigo 1.

    python3 validar.py                 # so o JSON
    python3 validar.py QT_teste.xlsx   # tambem as abas, exige recalculo antes
"""
import json, subprocess, sys, tempfile, pathlib, shutil

RAIZ = pathlib.Path(__file__).resolve().parent
BUILD = RAIZ / "build_planilha_escaneamentos.py"

ESPERADO_PERIODO = {
    "cliques brutos":        (["cliques_brutos"], 196),
    "descontados":           (["descontados"], 16),
    "escaneamentos limpos":  (["contados"], 180),
    "avaliacoes no periodo": (["periodo", "avaliacoes"], 126),
    "pares 1:1":             (["periodo", "viraram_nota"], 112),
    "orfas":                 (["periodo", "orfas"], 14),
    "escaneamentos perdidos":(["periodo", "perdidos"], 68),
    "conversao":             (["periodo", "conversao"], 0.622),
    "media das notas":       (["periodo", "media_notas"], 4.90),
}

# limpos, viraram nota, conversao, media, a pagar
ESPERADO_GARCOM = {
    "clara":     (104, 71, 0.683, 4.86, 192.0),
    "thalia":    (72,  41, 0.569, 4.98, 132.0),
    "alexandre": (2,    0, 0.0,   None,   0.0),
    "rafa":      (2,    0, 0.0,   None,   0.0),
}
TOTAL_A_PAGAR = 324.0


def cava(d, caminho):
    for k in caminho:
        d = d[k]
    return d


def perto(a, b, tol=0.005):
    if a is None or b is None:
        return a is b
    return abs(float(a) - float(b)) <= tol


def main():
    falhas = []
    trabalho = pathlib.Path(tempfile.mkdtemp(prefix="qt-validar-"))
    subprocess.run([sys.executable, str(RAIZ / "fixtures" / "gerar.py"), str(trabalho)],
                   check=True, capture_output=True)
    shutil.copy(BUILD, trabalho / BUILD.name)
    (trabalho / "perfil.json").write_text('{"nota_exibida": 4.2, "total_avaliacoes": 943}')
    saida = subprocess.run([sys.executable, BUILD.name, "QT_validacao.xlsx"],
                           cwd=trabalho, check=True, capture_output=True, text=True)
    j = json.loads(saida.stdout)

    print("Agosto de 2026, 05 a 24 — CLAUDE.md secao 8")
    print(f"{'medida':26} {'esperado':>10} {'obtido':>10}")
    for nome, (caminho, esp) in ESPERADO_PERIODO.items():
        obt = cava(j, caminho)
        ok = perto(obt, esp)
        if not ok: falhas.append(f"{nome}: esperado {esp}, obtido {obt}")
        print(f"{nome:26} {esp:>10} {obt:>10}  {'ok' if ok else 'DIVERGE'}")

    # por atendente, a partir do proprio pareamento
    import openpyxl
    wb = openpyxl.load_workbook(trabalho / "QT_validacao.xlsx")
    ws = wb["Pareamento"]
    notas, limpos = {}, {}
    for r in ws.iter_rows(min_row=6, values_only=True):
        if not isinstance(r[0], int): break
        notas.setdefault(r[1], []).append(r[6])
    ws = wb["Diario"]
    for r in ws.iter_rows(min_row=6, values_only=True):
        if not isinstance(r[3], int): break
        limpos[r[2]] = limpos.get(r[2], 0) + r[3]

    print()
    print(f"{'atendente':12} {'limpos':>7} {'notas':>6} {'conversao':>10} {'media':>7} {'a pagar':>9}")
    total_pagar = 0.0
    for g, (e_lim, e_not, e_conv, e_med, e_pag) in ESPERADO_GARCOM.items():
        lim = limpos.get(g, 0)
        ns = notas.get(g, [])
        conv = len(ns) / lim if lim else 0.0
        med = sum(ns) / len(ns) if ns else None
        pag = len(ns) * 2.0 + (len(ns) // 40) * 50.0
        total_pagar += pag
        ok = (lim == e_lim and len(ns) == e_not and perto(conv, e_conv)
              and perto(med, e_med, 0.005) and perto(pag, e_pag))
        if not ok:
            falhas.append(f"{g}: esperado {e_lim}/{e_not}/{e_conv}/{e_med}/{e_pag}, "
                          f"obtido {lim}/{len(ns)}/{round(conv,3)}/{med}/{pag}")
        m = "—" if med is None else f"{med:.2f}"
        print(f"{g:12} {lim:>7} {len(ns):>6} {conv:>9.1%} {m:>7} {pag:>8.0f}  {'ok' if ok else 'DIVERGE'}")
    ok = perto(total_pagar, TOTAL_A_PAGAR)
    if not ok: falhas.append(f"total a pagar: esperado {TOTAL_A_PAGAR}, obtido {total_pagar}")
    print(f"{'TOTAL':12} {'':>7} {'':>6} {'':>10} {'':>7} {total_pagar:>8.0f}  {'ok' if ok else 'DIVERGE'}")

    # formulas: confere que Matriz, Atendentes e Pagamento calculam mesmo.
    # O recalculo oficial e o LibreOffice; aqui e so uma conferencia independente,
    # e ela e opcional para nao virar dependencia da rodada.
    try:
        import formulas, logging
        logging.disable(logging.WARNING)
    except ImportError:
        print()
        print("formulas nao instalado, pulando a conferencia das formulas "
              "(pip install formulas)")
    else:
        print()
        xl = formulas.ExcelModel().loads(str(trabalho / "QT_validacao.xlsx")).finish()
        sol = xl.calculate()

        def celula(aba, ref):
            alvo = f"[QT_VALIDACAO.XLSX]{aba}'!{ref}".upper()
            for k in sol:
                if k.upper().endswith(alvo):
                    try: return sol[k].value[0, 0]
                    except Exception: return sol[k]
            return None

        # a linha TOTAL fica logo depois dos atendentes
        tot = 6 + len(ESPERADO_GARCOM)
        checa = [
            ("Atendentes escaneamentos", celula("ATENDENTES", f"B{tot}"), 180),
            ("Atendentes viraram nota",  celula("ATENDENTES", f"C{tot}"), 112),
            ("Atendentes conversao",     celula("ATENDENTES", f"E{tot}"), 0.622),
            ("Atendentes media",         celula("ATENDENTES", f"F{tot}"), 4.90),
            ("Atendentes participacao",  celula("ATENDENTES", f"I{tot}"), 1.0),
            ("Pagamento total a pagar",  celula("PAGAMENTO",  f"I{tot}"), 324.0),
            ("Pagamento custo/avaliacao",celula("PAGAMENTO",  f"C{tot+2}"), 2.89),
        ]
        for nome, obt, esp in checa:
            ok = perto(obt, esp, 0.005)
            if not ok: falhas.append(f"formula {nome}: esperado {esp}, obtido {obt}")
            o = "—" if obt is None else f"{float(obt):.3f}"
            print(f"formula {nome:28} {esp:>8} {o:>9}  {'ok' if ok else 'DIVERGE'}")

    # rodada degradada: sem avaliacoes.json o relatorio de escaneamentos tem que sair igual
    (trabalho / "avaliacoes.json").unlink()
    (trabalho / "avaliacoes_estado.json").unlink(missing_ok=True)
    deg = subprocess.run([sys.executable, BUILD.name, "QT_degradada.xlsx"],
                         cwd=trabalho, check=True, capture_output=True, text=True)
    d = json.loads(deg.stdout)
    print()
    checks = [
        ("escaneamentos preservados", d["contados"] == 180),
        ("pares zerados", d["periodo"]["viraram_nota"] == 0),
        ("avaliacoes_ausentes", d["relatorio"]["avaliacoes_ausentes"] is True),
        ("planilha gerada", (trabalho / "QT_degradada.xlsx").exists()),
    ]
    for nome, cond in checks:
        if not cond: falhas.append(f"rodada degradada, {nome}")
        print(f"rodada degradada, {nome:28} {'ok' if cond else 'DIVERGE'}")

    print()
    if falhas:
        print(f"{len(falhas)} divergencia(s):")
        for f in falhas: print("  -", f)
        return 1
    print("Tudo bate com os numeros medidos.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
